#!/usr/bin/env python3
"""
Monitora uma pasta por alterações em vendas.csv ou novos arquivos Excel (.xlsx/.xls).
Ao detectar novas linhas, extrai produto, valor e vendedor e envia
mensagem via Evolution API (WhatsApp).
"""

from __future__ import annotations

import csv
import json
import logging
import os
import sys
import time
from pathlib import Path
from typing import Any

import requests
from dotenv import load_dotenv
from openpyxl import load_workbook
from watchdog.events import FileSystemEventHandler
from watchdog.observers import Observer

load_dotenv()

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
)
logger = logging.getLogger("monitor_vendas")

# ---------------------------------------------------------------------------
# Configuração
# ---------------------------------------------------------------------------

MONITOR_DIR = Path(os.getenv("MONITOR_DIR", "./dados")).resolve()
VENDAS_CSV = "vendas.csv"
EXCEL_EXTENSIONS = {".xlsx", ".xls"}

EVOLUTION_API_URL = os.getenv("EVOLUTION_API_URL", "").rstrip("/")
EVOLUTION_API_KEY = os.getenv("EVOLUTION_API_KEY", "")
EVOLUTION_INSTANCE = os.getenv("EVOLUTION_INSTANCE", "")
EVOLUTION_WEBHOOK_URL = os.getenv("EVOLUTION_WEBHOOK_URL", "").strip()
WHATSAPP_NUMBER = os.getenv("WHATSAPP_NUMBER", "")

# Colunas esperadas (nomes normalizados -> chave interna)
COLUMN_ALIASES = {
    "produto": "produto",
    "product": "produto",
    "valor": "valor",
    "value": "valor",
    "preco": "valor",
    "preço": "valor",
    "vendedor": "vendedor",
    "seller": "vendedor",
}

REQUIRED_FIELDS = ("produto", "valor", "vendedor")

# Debounce: evita processar o mesmo arquivo várias vezes em milissegundos
DEBOUNCE_SECONDS = 1.0


def _normalize_header(name: str) -> str | None:
    key = (name or "").strip().lower()
    return COLUMN_ALIASES.get(key)


def _row_to_sale(raw: dict[str, Any]) -> dict[str, str] | None:
    """Mapeia um dict de colunas brutas para {produto, valor, vendedor}."""
    mapped: dict[str, str] = {}
    for col, value in raw.items():
        field = _normalize_header(str(col) if col is not None else "")
        if field and field not in mapped:
            mapped[field] = "" if value is None else str(value).strip()

    if not all(mapped.get(f) for f in REQUIRED_FIELDS):
        return None
    return {f: mapped[f] for f in REQUIRED_FIELDS}


def format_message(sale: dict[str, str]) -> str:
    return (
        "🛒 *Nova venda registrada!*\n\n"
        f"*Produto:* {sale['produto']}\n"
        f"*Valor:* {sale['valor']}\n"
        f"*Vendedor:* {sale['vendedor']}"
    )


DRY_RUN = os.getenv("DRY_RUN", "").lower() in {"1", "true", "yes"}


def send_whatsapp(sale: dict[str, str]) -> bool:
    """Envia mensagem de texto via Evolution API / webhook."""
    text = format_message(sale)

    if DRY_RUN:
        logger.info("[DRY-RUN] Mensagem para %s:\n%s", WHATSAPP_NUMBER or "(não definido)", text)
        return True

    if not WHATSAPP_NUMBER:
        logger.error("WHATSAPP_NUMBER não configurado.")
        return False

    url = EVOLUTION_WEBHOOK_URL
    if not url:
        if not (EVOLUTION_API_URL and EVOLUTION_INSTANCE):
            logger.error(
                "Configure EVOLUTION_WEBHOOK_URL ou "
                "EVOLUTION_API_URL + EVOLUTION_INSTANCE."
            )
            return False
        url = f"{EVOLUTION_API_URL}/message/sendText/{EVOLUTION_INSTANCE}"

    payload = {
        "number": WHATSAPP_NUMBER,
        "text": text,
    }
    headers = {
        "Content-Type": "application/json",
        "apikey": EVOLUTION_API_KEY,
    }

    try:
        response = requests.post(url, json=payload, headers=headers, timeout=30)
        response.raise_for_status()
        logger.info(
            "Mensagem enviada para %s | produto=%s",
            WHATSAPP_NUMBER,
            sale["produto"],
        )
        return True
    except requests.RequestException as exc:
        logger.error("Falha ao enviar WhatsApp: %s", exc)
        if getattr(exc, "response", None) is not None:
            logger.error("Resposta: %s", exc.response.text[:500])
        return False


# ---------------------------------------------------------------------------
# Leitura de arquivos
# ---------------------------------------------------------------------------


def read_csv_rows(path: Path) -> list[dict[str, str]]:
    sales: list[dict[str, str]] = []
    with path.open("r", encoding="utf-8-sig", newline="") as fh:
        # Tenta detectar delimitador (vírgula ou ponto-e-vírgula)
        sample = fh.read(4096)
        fh.seek(0)
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=",;")
        except csv.Error:
            dialect = csv.excel

        reader = csv.DictReader(fh, dialect=dialect)
        for raw in reader:
            sale = _row_to_sale(raw)
            if sale:
                sales.append(sale)
    return sales


def read_excel_rows(path: Path) -> list[dict[str, str]]:
    sales: list[dict[str, str]] = []
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb.active
        rows = ws.iter_rows(values_only=True)
        try:
            header = next(rows)
        except StopIteration:
            return sales

        headers = [str(h).strip() if h is not None else "" for h in header]
        for values in rows:
            if values is None or all(v is None or str(v).strip() == "" for v in values):
                continue
            raw = {headers[i]: values[i] for i in range(min(len(headers), len(values)))}
            sale = _row_to_sale(raw)
            if sale:
                sales.append(sale)
    finally:
        wb.close()
    return sales


def sale_fingerprint(sale: dict[str, str]) -> str:
    return json.dumps(sale, ensure_ascii=False, sort_keys=True)


# ---------------------------------------------------------------------------
# Estado e handler
# ---------------------------------------------------------------------------


class SalesMonitor(FileSystemEventHandler):
    """Observa vendas.csv e novos/alterados arquivos Excel."""

    def __init__(self, watch_dir: Path) -> None:
        super().__init__()
        self.watch_dir = watch_dir
        self._last_processed: dict[str, float] = {}
        # fingerprints já notificados por arquivo
        self._seen: dict[str, set[str]] = {}
        self._bootstrap()

    def _bootstrap(self) -> None:
        """Carrega estado inicial sem enviar mensagens (somente baseline)."""
        csv_path = self.watch_dir / VENDAS_CSV
        if csv_path.is_file():
            self._seen[str(csv_path)] = {
                sale_fingerprint(s) for s in read_csv_rows(csv_path)
            }
            logger.info(
                "Baseline CSV: %d linha(s) em %s",
                len(self._seen[str(csv_path)]),
                csv_path.name,
            )

        for path in self.watch_dir.iterdir():
            if path.is_file() and path.suffix.lower() in EXCEL_EXTENSIONS:
                try:
                    rows = read_excel_rows(path)
                except Exception as exc:  # noqa: BLE001
                    logger.warning("Ignorando Excel no bootstrap (%s): %s", path.name, exc)
                    continue
                self._seen[str(path)] = {sale_fingerprint(s) for s in rows}
                logger.info(
                    "Baseline Excel: %d linha(s) em %s",
                    len(self._seen[str(path)]),
                    path.name,
                )

    def _should_debounce(self, path: Path) -> bool:
        key = str(path)
        now = time.monotonic()
        last = self._last_processed.get(key, 0.0)
        if now - last < DEBOUNCE_SECONDS:
            return True
        self._last_processed[key] = now
        return False

    def _wait_stable(self, path: Path, retries: int = 5, delay: float = 0.3) -> bool:
        """Espera o arquivo terminar de ser escrito (tamanho estável)."""
        prev = -1
        for _ in range(retries):
            try:
                size = path.stat().st_size
            except OSError:
                time.sleep(delay)
                continue
            if size == prev and size > 0:
                return True
            prev = size
            time.sleep(delay)
        return path.exists()

    def _process_file(self, path: Path) -> None:
        if not path.is_file():
            return
        if self._should_debounce(path):
            return
        if not self._wait_stable(path):
            logger.warning("Arquivo instável, pulando: %s", path.name)
            return

        name = path.name.lower()
        suffix = path.suffix.lower()

        try:
            if name == VENDAS_CSV:
                sales = read_csv_rows(path)
            elif suffix in EXCEL_EXTENSIONS:
                sales = read_excel_rows(path)
            else:
                return
        except Exception as exc:  # noqa: BLE001
            logger.error("Erro ao ler %s: %s", path.name, exc)
            return

        key = str(path)
        # Arquivo ainda não visto (ex.: Excel novo): known vazio → todas as linhas são novas
        known = self._seen.setdefault(key, set())

        new_sales = [s for s in sales if sale_fingerprint(s) not in known]
        if not new_sales:
            logger.debug("Nenhuma linha nova em %s", path.name)
            return

        logger.info("%d nova(s) venda(s) em %s", len(new_sales), path.name)
        for sale in new_sales:
            if send_whatsapp(sale):
                known.add(sale_fingerprint(sale))
            else:
                logger.warning("Linha não marcada como vista (falha no envio): %s", sale)

        self._seen[key] = known

    def on_created(self, event) -> None:  # noqa: ANN001
        if event.is_directory:
            return
        path = Path(event.src_path)
        if path.name.startswith(".") or path.name.endswith("~"):
            return

        name = path.name.lower()
        suffix = path.suffix.lower()
        if name != VENDAS_CSV and suffix not in EXCEL_EXTENSIONS:
            return

        logger.info("Arquivo criado: %s", path.name)
        self._process_file(path)

    def on_modified(self, event) -> None:  # noqa: ANN001
        if event.is_directory:
            return
        path = Path(event.src_path)
        if path.name.startswith(".") or path.name.endswith("~"):
            return

        name = path.name.lower()
        suffix = path.suffix.lower()
        if name != VENDAS_CSV and suffix not in EXCEL_EXTENSIONS:
            return

        logger.info("Arquivo modificado: %s", path.name)
        self._process_file(path)


def validate_config() -> None:
    if DRY_RUN:
        logger.warning("Modo DRY-RUN ativo: mensagens não serão enviadas.")
        return

    missing = []
    if not WHATSAPP_NUMBER:
        missing.append("WHATSAPP_NUMBER")
    if not EVOLUTION_WEBHOOK_URL:
        if not EVOLUTION_API_URL:
            missing.append("EVOLUTION_API_URL (ou EVOLUTION_WEBHOOK_URL)")
        if not EVOLUTION_INSTANCE:
            missing.append("EVOLUTION_INSTANCE (ou EVOLUTION_WEBHOOK_URL)")
    if not EVOLUTION_API_KEY:
        missing.append("EVOLUTION_API_KEY")
    if missing:
        logger.error("Variáveis obrigatórias ausentes: %s", ", ".join(missing))
        logger.error("Copie .env.example para .env e preencha os valores.")
        sys.exit(1)


def main() -> None:
    validate_config()
    MONITOR_DIR.mkdir(parents=True, exist_ok=True)

    handler = SalesMonitor(MONITOR_DIR)
    observer = Observer()
    observer.schedule(handler, str(MONITOR_DIR), recursive=False)
    observer.start()

    logger.info("Monitorando pasta: %s", MONITOR_DIR)
    logger.info("Arquivos: %s | Excel: %s", VENDAS_CSV, ", ".join(sorted(EXCEL_EXTENSIONS)))
    logger.info("WhatsApp destino: %s", WHATSAPP_NUMBER or "(dry-run)")
    logger.info("Pressione Ctrl+C para encerrar.")

    try:
        while True:
            time.sleep(1)
    except KeyboardInterrupt:
        logger.info("Encerrando...")
    finally:
        observer.stop()
        observer.join()


if __name__ == "__main__":
    main()
