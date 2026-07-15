"""Leitura de arquivos de vendas em CSV ou Excel (.xlsx).

O parser converte cada linha em um dicionário e normaliza o campo de ID para
inteiro, descartando linhas sem ID válido.
"""
from __future__ import annotations

import csv
import os
from typing import Any


class SalesParser:
    """Lê registros de vendas de um arquivo CSV ou Excel."""

    def __init__(self, file_path: str, id_field: str = "id"):
        self.file_path = file_path
        self.id_field = id_field

    def read(self) -> list[dict[str, Any]]:
        """Retorna a lista de vendas presentes no arquivo.

        Retorna lista vazia se o arquivo ainda não existir.
        """
        if not os.path.exists(self.file_path):
            return []

        ext = os.path.splitext(self.file_path)[1].lower()
        if ext in (".xlsx", ".xlsm", ".xls"):
            return self._read_excel()
        return self._read_csv()

    def _read_csv(self) -> list[dict[str, Any]]:
        records: list[dict[str, Any]] = []
        # utf-8-sig lida com BOM gerado por Excel/Windows.
        with open(self.file_path, newline="", encoding="utf-8-sig") as handle:
            reader = csv.DictReader(handle)
            for row in reader:
                record = self._normalize(row)
                if record is not None:
                    records.append(record)
        return records

    def _read_excel(self) -> list[dict[str, Any]]:
        from openpyxl import load_workbook

        workbook = load_workbook(self.file_path, read_only=True, data_only=True)
        try:
            worksheet = workbook.active
            rows = worksheet.iter_rows(values_only=True)
            try:
                header = [
                    str(cell).strip() if cell is not None else ""
                    for cell in next(rows)
                ]
            except StopIteration:
                return []

            records: list[dict[str, Any]] = []
            for values in rows:
                row = {
                    header[i]: values[i]
                    for i in range(min(len(header), len(values)))
                }
                record = self._normalize(row)
                if record is not None:
                    records.append(record)
            return records
        finally:
            workbook.close()

    def _normalize(self, row: dict[str, Any]) -> dict[str, Any] | None:
        raw_id = row.get(self.id_field)
        if raw_id is None or str(raw_id).strip() == "":
            return None
        try:
            sale_id = int(float(str(raw_id).strip()))
        except (ValueError, TypeError):
            return None

        clean = {
            key: (value.strip() if isinstance(value, str) else value)
            for key, value in row.items()
        }
        clean[self.id_field] = sale_id
        return clean
